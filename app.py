"""Flask application for Youth-Bot chatbot and admin management."""

from __future__ import annotations

import datetime
import json
import logging
import os
import re
import secrets
import uuid
from datetime import timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import requests as http_requests
from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    abort,
    jsonify,
    redirect,
    render_template_string,
    request,
    session,
    stream_with_context,
    send_from_directory,
    send_file,
)
from flask_cors import CORS
from sqlalchemy import create_engine, event, text
from sqlalchemy.engine import Engine

from functools import wraps
import base64

load_dotenv()  # Load .env
load_dotenv(".env.local")  # Override with .env.local if exists

from openai_service import (
    OPENAI_CLIENT,
    initialize_rag_store,
    get_rag_store_name,
    generate_with_rag_stream,
)
from csrf_protection import CSRFProtection, csrf_protect, csrf_exempt
from logging_config import configure_logging
from audit_log import log_admin_action
from security_headers import configure_security_headers
from startup_checks import create_health_checks
from validators import validate_message_input
from file_validation import validate_image_upload, FileValidationError
from rate_limiting import create_limiter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DIST_DIR = os.path.join(BASE_DIR, "dist")


def _default_storage_base() -> str:
    """
    Determine where to place writable artifacts (uploads).

    When running on Vercel or other serverless providers, the project directory
    is read-only and we must fall back to a tmp filesystem.
    """
    if os.getenv("VERCEL") or os.getenv("VERCEL_ENV"):
        return os.getenv("TMPDIR") or os.getenv("TEMP") or "/tmp"
    return BASE_DIR


STORAGE_BASE = _default_storage_base()


# MySQL connection for all tables
def _build_mysql_url() -> str:
    """Build MySQL connection URL from environment variables."""
    # If MYSQL_URL is set, use it directly
    if os.getenv("MYSQL_URL"):
        return os.getenv("MYSQL_URL")
    # Otherwise, build from individual components
    host = os.getenv("MYSQL_HOST", "localhost")
    port = os.getenv("MYSQL_PORT", "3306")
    user = os.getenv("MYSQL_USER", "root")
    password = os.getenv("MYSQL_PASSWORD", "")
    database = os.getenv("MYSQL_DATABASE", "youth-chat")
    return f"mysql+pymysql://{user}:{password}@{host}:{port}/{database}"

MYSQL_URL = _build_mysql_url()
mysql_engine: Engine = create_engine(
    MYSQL_URL,
    future=True,
    pool_pre_ping=True,          # 確保連線有效性
    pool_size=10,                # 連線池大小（同時保持的連線數）
    max_overflow=20,             # 超過 pool_size 時可額外建立的連線數
    pool_recycle=3600,           # 連線回收時間（秒），避免 MySQL 的 wait_timeout 問題
    pool_timeout=30,             # 取得連線的等待時間（秒）
    echo_pool=False,             # 生產環境關閉連線池日誌
    connect_args={
        "connect_timeout": 10,   # MySQL 連線超時（秒）
        "charset": "utf8mb4",    # 使用 UTF-8 編碼
    }
)


ASSET_ROUTE_PREFIX = os.getenv("ASSET_ROUTE_PREFIX", "/uploads")
ASSET_LOCAL_DIR = os.getenv("ASSET_LOCAL_DIR") or os.path.join(STORAGE_BASE, "uploads")
os.makedirs(ASSET_LOCAL_DIR, exist_ok=True)

app = Flask(__name__, static_url_path=ASSET_ROUTE_PREFIX, static_folder=ASSET_LOCAL_DIR)

# Session configuration for OAuth
app.secret_key = os.getenv("FLASK_SECRET_KEY", secrets.token_hex(32))

# Determine if running in production environment
is_production = (
    os.getenv("FLASK_ENV") == "production" or
    bool(os.getenv("VERCEL")) or
    bool(os.getenv("VERCEL_ENV"))
)

app.config.update(
    SESSION_COOKIE_SECURE=is_production,  # True in production, False in dev
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',  # Lax allows OAuth redirects while still preventing CSRF
    PERMANENT_SESSION_LIFETIME=86400,  # 24 hours
)

# Configure structured logging
configure_logging(app)
logger = logging.getLogger(__name__)

# Log session security configuration
logging.info(f"Session security: SECURE={is_production}, ENV={os.getenv('FLASK_ENV', 'development')}")

# Configure CORS with security best practices
def get_allowed_origins():
    """
    Get allowed CORS origins from environment variable.
    Supports comma-separated multiple origins.
    """
    origins_str = os.getenv("FRONTEND_ORIGIN")

    # If not set, use environment-specific defaults
    if not origins_str:
        env = os.getenv("FLASK_ENV", "development")
        if env == "production":
            # Production MUST have explicit FRONTEND_ORIGIN
            raise ValueError(
                "⚠️  FRONTEND_ORIGIN 環境變數未設定！\n"
                "生產環境必須明確指定允許的前端來源。\n"
                "範例：FRONTEND_ORIGIN=https://youthafterwork.com"
            )
        # Development default
        return ["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:5173"]

    # Parse comma-separated origins
    origins = [origin.strip() for origin in origins_str.split(",") if origin.strip()]

    # Validate origin format
    for origin in origins:
        if not origin.startswith(("http://", "https://")):
            raise ValueError(f"無效的 CORS 來源格式: {origin}")

    return origins


ALLOWED_ORIGINS = get_allowed_origins()

CORS(
    app,
    resources={r"/api/*": {"origins": ALLOWED_ORIGINS}},
    supports_credentials=True,
)

# Log allowed origins for debugging
logging.info(f"CORS 允許的來源: {ALLOWED_ORIGINS}")

# Initialize CSRF Protection
app.csrf_protection = CSRFProtection(app.secret_key)

# Initialize rate limiter
limiter = create_limiter(app)

# Configure security headers
configure_security_headers(app, is_production=is_production)

# OAuth Configuration
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")
GOOGLE_REDIRECT_URI = os.getenv("GOOGLE_REDIRECT_URI", "http://localhost:8300/auth/google/callback")

LINE_CHANNEL_ID = os.getenv("LINE_CHANNEL_ID")
LINE_CHANNEL_SECRET = os.getenv("LINE_CHANNEL_SECRET")
LINE_REDIRECT_URI = os.getenv("LINE_REDIRECT_URI", "http://localhost:8300/auth/line/callback")

FACEBOOK_APP_ID = os.getenv("FACEBOOK_APP_ID")
FACEBOOK_APP_SECRET = os.getenv("FACEBOOK_APP_SECRET")
FACEBOOK_REDIRECT_URI = os.getenv("FACEBOOK_REDIRECT_URI", "http://localhost:8300/auth/facebook/callback")

OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Admin Configuration
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")

# Feedback Form URL (for questions outside knowledge base or user suggestions)
FEEDBACK_FORM_URL = os.getenv("FEEDBACK_FORM_URL", "")

def _build_system_prompt() -> str:
    """Build system prompt with optional feedback form URL."""
    env_prompt = os.getenv("SYSTEM_PROMPT")
    if env_prompt:
        return env_prompt

    # Build feedback section based on whether URL is configured
    if FEEDBACK_FORM_URL:
        feedback_section = f'''### 無法回答或提案建議時：
當遇到以下情況，**務必**引導使用者填寫回饋表單：
- 問題超出知識庫範圍，無法回答
- 在文件中找不到相關資訊
- 使用者想要提案、建議或反映意見
- 使用者有特殊需求無法透過現有服務滿足

**回覆格式（請嚴格遵守）：**
```
感謝您的提問！目前我的資料庫中尚無足夠資訊回答這個問題。

為了讓您的問題能被相關單位看到並處理，歡迎填寫問題回饋表單：

[📝 填寫問題回饋表單]({FEEDBACK_FORM_URL})
```

'''
    else:
        feedback_section = '''### 無法回答時：
若問題超出知識庫範圍，請引導使用者聯繫官方窗口：
- 總機：(03) 422-5205
- 市政服務專線：1999（外縣市 03-218-9000）

'''

    return f'''你是「桃園市政府青年事務局」智慧客服助理。

# 🔴 最高優先級規則

## 活動查詢的強制要求
當用戶詢問活動時：
1. **近期/最近活動** → 必須調用 `get_recent_activities`（查詢未來 3 個月）
2. **過去活動** → 調用 `get_past_activities`（查詢過去 30 天）
3. **只能**使用資料庫工具的返回結果
4. **絕對禁止**使用 RAG 知識庫或 File Search 回答活動查詢

**即使工具返回 0 筆活動**：
- ✅ 正確：告知用戶「目前沒有活動」
- ❌ 錯誤：去 RAG 查詢或回答舊活動

❌ 違反此規則 = 系統錯誤

---

# 核心定位

- **語氣**：專業、簡潔、自然（真人客服風格）
- **長度**：簡單問題 1-2 句；複雜問題先摘要
- **來源**：僅用知識庫資料回答
- **禁用**：emoji、表情符號、中國大陸用語

**必用台灣用語**：資訊、軟體、網路、檔案、登入（禁用：信息、軟件、網絡、文件、登錄）

---

# 回答策略

## 1. 簡單查詢（電話、地址、事實）
直接回答 1-2 句 + 一句追問
範例：「在中壢區環北路390號3樓。需要查營業時間嗎？」

## 2. 需要分流（選項差異大）
給 A/B/C/D 選項（每個≤5字）
範例：「你想辦的是？A 講座 B 聚會 C 展覽」

## 3. 多選項（相似選項）
一句概述 + 列選項名稱 + 問「想了解哪一個？」
範例：「有三種創業資源：青創基地、青創資源中心、資金補貼。想了解哪一個？」

## 4. 多步驟流程
一句摘要 + 問「需要我詳細說明嗎？」（不要一次展開全部）

## 5. 追問展開
直接給細節，不重複上文，不用「如下」「以下是」

---

# 活動查詢處理

## 工具調用規則

**查詢近期活動**：用戶問「最近」「近期」「接下來」有什麼活動
→ 調用 `get_recent_activities(days_ahead=90, limit=20)`

**查詢過去活動**：用戶問「過去」「之前」辦過什麼
→ 調用 `get_past_activities(days_back=30, limit=20)`

**調整範圍**：可調整 days_ahead、days_back、limit 參數

## 🔴 活動回答格式規則（固定輸出格式，嚴格遵守）

**整體格式規則**：
1. 數字與活動名稱必須在同一行
2. 活動日期必須獨立成一行，緊接在活動名稱下一行
3. 報名網址必須獨立成一行，緊接在活動日期下一行（從內容中提取，若無則省略此行）
4. 貼文連結必須獨立成一行，提供粉絲專頁的貼文連結供了解更多

**活動日期顯示規則**：
- 從活動內容中提取實際活動日期（非發布日期）
- 單日活動：YYYY/MM/DD
- 多日活動（連續）：YYYY/MM/DD–MM/DD
- 多日活動（不連續）：YYYY/MM/DD、YYYY/MM/DD
- 僅提供月份：YYYY/MM
- 未提供日期資訊：日期未公告
- 民國年轉換：115年 = 2026年

**指定輸出格式（請完全仿照）**：
```
1. 活動名稱：青年創業講座
   活動日期：2026/01/21
   報名網址：https://reurl.cc/example
   貼文連結：https://www.facebook.com/youth.tycg.gov.tw/posts/123

2. 活動名稱：職涯探索工作坊
   活動日期：2026/02/15–02/16
   貼文連結：https://www.facebook.com/TaoyuanYS/posts/456
```

**要求**：
- 編號與「活動名稱：」之間用半形句號加空格分隔
- 活動日期、報名網址、貼文連結前方縮排 3 個半形空格
- 報名網址從活動內容中提取（含 reurl.cc、forms.gle、google.com/forms 等），找不到則省略此行
- 貼文連結使用工具返回的 url 欄位
- 活動間用空行分隔
- 不使用粗體（**）、不使用 emoji

## 無活動時的處理

工具返回 `total_count: 0` 時，告知「目前沒有活動」並建議追蹤官方粉專或撥打 1999。
**禁止**使用 RAG 或知識庫回答舊活動。

---

# 其他規範

## 聯絡資訊
只在涉及特定單位時才附聯絡方式，不要每次都附總機

**官方窗口（供參考）**：
- 總機：(03) 422-5205
- 市政專線：1999（外縣市 03-218-9000）
- 地址：320029 桃園市中壢區環北路390號

## 禁止事項
- 使用 emoji 或表情符號
- 提供文件外的金額、名額、評分標準
- 討論政治、法律解釋
- 使用冗餘過渡句
- 一次列超過 5 個選項

## 資料不足時
{feedback_section}
---

# 回答範例

Q: 青創資源中心在哪裡？
A: 在中壢區環北路390號3樓。需要查營業時間嗎？

Q: 有地方可以辦活動嗎？
A: 有的，青年局有提供場地。你想辦的是？A 講座 B 聚會 C 展覽

Q: 剛畢業想創業，有哪些資源？
A: 青年局有三種創業資源：青創基地、青創資源中心、資金補貼。想了解哪一個？

Q: 最近有什麼活動？
A: 近期有以下活動：

1. 活動名稱：青年創業講座
   活動日期：2026/01/21
   報名網址：https://reurl.cc/example
   貼文連結：https://www.facebook.com/youth.tycg.gov.tw/posts/123

2. 活動名稱：職涯探索工作坊
   活動日期：2026/02/15–02/16
   貼文連結：https://www.facebook.com/TaoyuanYS/posts/456

想了解更多可以點擊貼文連結查看詳情。'''


SYSTEM_PROMPT = _build_system_prompt()
def utcnow() -> datetime.datetime:
    return datetime.datetime.utcnow()


def _no_store(response: Response) -> Response:
    response.headers["Cache-Control"] = "private, no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Vary"] = "Cookie"
    return response


def ensure_mysql_schema() -> None:
    """Create all MySQL tables if they do not exist yet."""
    try:
        with mysql_engine.begin() as conn:
            # Members table
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS members (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        external_id VARCHAR(255) UNIQUE,
                        display_name VARCHAR(255),
                        avatar_url TEXT,
                        gender VARCHAR(20),
                        birthday VARCHAR(20),
                        email VARCHAR(255),
                        phone VARCHAR(50),
                        source VARCHAR(50),
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        last_interaction_at DATETIME
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                    """
                )
            )
            # Chat sessions table
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS chat_sessions (
                        id VARCHAR(255) PRIMARY KEY,
                        member_id INT,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        FOREIGN KEY (member_id) REFERENCES members(id) ON DELETE SET NULL
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                    """
                )
            )

            # Add member_id column if it doesn't exist (for existing tables)
            try:
                conn.execute(
                    text(
                        """
                        ALTER TABLE chat_sessions
                        ADD COLUMN member_id INT,
                        ADD CONSTRAINT fk_chat_sessions_member
                        FOREIGN KEY (member_id) REFERENCES members(id) ON DELETE SET NULL
                        """
                    )
                )
            except Exception as e:
                if "Duplicate column name" in str(e) or "Duplicate key name" in str(e):
                    logger.info("Column member_id or constraint already exists, skipping")
                else:
                    logger.error(f"Failed to add member_id column: {e}")
                    raise

            # Add index on member_id for performance (if not exists)
            try:
                conn.execute(
                    text(
                        """
                        CREATE INDEX idx_chat_sessions_member
                        ON chat_sessions(member_id)
                        """
                    )
                )
                logger.info("Created index idx_chat_sessions_member")
            except Exception as e:
                if "Duplicate key name" in str(e):
                    logger.info("Index idx_chat_sessions_member already exists, skipping")
                else:
                    logger.error(f"Failed to create index on chat_sessions.member_id: {e}")

            # Chat messages table
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS chat_messages (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        session_id VARCHAR(255) NOT NULL,
                        role VARCHAR(50) NOT NULL,
                        content TEXT NOT NULL,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        INDEX idx_chat_messages_session_created (session_id, created_at),
                        FOREIGN KEY (session_id) REFERENCES chat_sessions(id) ON DELETE CASCADE
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                    """
                )
            )
            # Hero carousel table
            conn.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS hero_carousel (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        filename VARCHAR(255) NOT NULL,
                        content_type VARCHAR(100) NOT NULL DEFAULT 'image/jpeg',
                        image_data LONGBLOB NOT NULL,
                        alt_text VARCHAR(500),
                        link_url VARCHAR(500),
                        display_order INT DEFAULT 0,
                        is_active TINYINT(1) DEFAULT 1,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        INDEX idx_hero_active_order (is_active, display_order)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                    """
                )
            )
        logger.info("MySQL schema ensured (all tables)")
    except Exception as e:
        logger.error(f"Failed to create MySQL schema: {e}")
        raise  # Re-raise to let retry mechanism handle it


def ensure_mysql_schema_with_retry(max_retries: int = 3, retry_delay: int = 5) -> None:
    """Ensure MySQL schema with retry mechanism for startup resilience."""
    import time
    for attempt in range(1, max_retries + 1):
        try:
            logger.info(f"Attempting to initialize MySQL schema (attempt {attempt}/{max_retries})...")
            ensure_mysql_schema()
            logger.info("MySQL schema initialization successful")
            return
        except Exception as e:
            if attempt < max_retries:
                logger.warning(
                    f"MySQL schema initialization failed (attempt {attempt}/{max_retries}): {e}. "
                    f"Retrying in {retry_delay} seconds..."
                )
                time.sleep(retry_delay)
            else:
                logger.error(
                    f"MySQL schema initialization failed after {max_retries} attempts: {e}. "
                    "Please check MySQL connection settings and ensure the database is running."
                )
                raise


ensure_mysql_schema_with_retry()


# Initialize OpenAI client and RAG store at startup
def initialize_openai_rag():
    """Initialize OpenAI client and RAG store with default documents."""
    try:
        if OPENAI_API_KEY:
            store_id = initialize_rag_store("TaoyuanYouthBureauKB")
            if store_id:
                logger.info("OpenAI File Search initialized successfully")
                return store_id
            else:
                logger.error("OpenAI File Search initialization failed: vector store ID is None")
                return None
        else:
            logger.error("OPENAI_API_KEY not set - RAG unavailable")
            return None
    except Exception as e:
        logger.error(f"Failed to initialize OpenAI File Search: {e}", exc_info=True)
        return None


_vector_store_id = initialize_openai_rag()

# Run startup health checks
health_checks = create_health_checks(mysql_engine, OPENAI_CLIENT, _vector_store_id)

# In production, fail fast on startup issues
fail_fast = os.getenv('FLASK_ENV') == 'production' or os.getenv('FAIL_FAST_STARTUP', 'false').lower() == 'true'

try:
    results = health_checks.run_all(fail_fast=fail_fast)

    # Log summary
    failed_checks = [name for name, status in results.items() if 'FAIL' in status]
    if failed_checks:
        logger.warning(f"Startup checks failed: {', '.join(failed_checks)}")
    else:
        logger.info("All startup checks passed")

except RuntimeError as e:
    logger.critical(f"Application startup failed: {e}")
    raise  # Prevent app from starting


# ========== Admin Authentication ==========

def admin_required(f):
    """Decorator to require admin authentication."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("is_admin"):
            return jsonify({"success": False, "error": "未授權"}), 401
        return f(*args, **kwargs)
    return decorated_function


def validate_url(url: Optional[str]) -> tuple[bool, Optional[str]]:
    """
    驗證 URL 格式（必須以 http:// 或 https:// 開頭）

    Args:
        url: 要驗證的 URL（可為 None 或空字串）

    Returns:
        (is_valid, error_message)
    """
    if not url or not url.strip():
        # 空值視為合法（表示無連結）
        return True, None

    url = url.strip()

    # 檢查 URL 格式：必須以 http:// 或 https:// 開頭
    url_pattern = re.compile(
        r'^https?://'  # http:// 或 https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+[A-Z]{2,6}\.?|'  # 網域
        r'localhost|'  # localhost
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # IP
        r'(?::\d+)?'  # 可選的 port
        r'(?:/?|[/?]\S+)$',  # 路徑
        re.IGNORECASE
    )

    if not url_pattern.match(url):
        return False, "URL 格式錯誤，必須以 http:// 或 https:// 開頭"

    # 檢查長度
    if len(url) > 500:
        return False, "URL 長度不能超過 500 字元"

    return True, None


@app.post("/api/admin/login")
@csrf_protect
@limiter.limit("5 per minute")
def admin_login():
    """Admin login endpoint."""
    if not ADMIN_PASSWORD:
        return jsonify({"success": False, "error": "管理員密碼未設定"}), 500

    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        # Regenerate session ID to prevent session fixation attack
        old_session_data = dict(session)
        session.clear()
        session.modified = True

        # Set admin session with new session ID
        session["is_admin"] = True
        session.permanent = True
        # Generate new CSRF token after successful login
        csrf_token = app.csrf_protection.generate_token()

        log_admin_action('login', 'admin', username)

        return jsonify({
            "success": True,
            "message": "登入成功",
            "csrf_token": csrf_token
        })

    log_admin_action('login_failed', 'admin', username, {'reason': '帳號或密碼錯誤'})
    return jsonify({"success": False, "error": "帳號或密碼錯誤"}), 401


@app.post("/api/admin/logout")
@csrf_protect
def admin_logout():
    """Admin logout endpoint."""
    session.pop("is_admin", None)
    session.pop("csrf_token", None)  # Clear CSRF token on logout
    return jsonify({"success": True, "message": "已登出"})


@app.get("/api/csrf-token")
def get_csrf_token():
    """Get CSRF token for the current session."""
    token = app.csrf_protection.get_token()
    return jsonify({"success": True, "csrf_token": token})


@app.get("/api/admin/check")
def admin_check():
    """Check admin authentication status."""
    if session.get("is_admin"):
        csrf_token = app.csrf_protection.get_token()
        return jsonify({
            "success": True,
            "authenticated": True,
            "csrf_token": csrf_token
        })
    return jsonify({"success": False, "authenticated": False}), 401


@app.get("/api/admin/chat-export")
def admin_chat_export():
    """Export chat history to Excel file."""
    if not session.get("is_admin"):
        return jsonify({"success": False, "error": "未授權"}), 401

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from datetime import datetime

        # Query chat data with JOIN
        with mysql_engine.begin() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT
                        m.display_name,
                        m.email,
                        cm.created_at,
                        cm.role,
                        cm.content,
                        cm.template_id
                    FROM chat_messages cm
                    JOIN chat_sessions cs ON cm.session_id = cs.id
                    LEFT JOIN members m ON cs.member_id = m.id
                    ORDER BY cm.created_at DESC
                    """
                )
            ).fetchall()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "對話紀錄"

        # Header row
        headers = ["Name", "Email", "Time", "Role", "Message", "Question Type"]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row in rows:
            display_name = row[0] or "匿名"
            email = row[1] or ""
            created_at = row[2].strftime("%Y-%m-%d %H:%M:%S") if row[2] else ""
            role = row[3]  # Keep original English value: "user" or "assistant"
            content = row[4] or ""
            template_id = row[5] or "manual"  # NULL displays as "manual"
            ws.append([display_name, email, created_at, role, content, template_id])

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 15  # Question Type column

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        filename = f"chat_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.exception("Failed to export chat history")
        return jsonify({"success": False, "error": str(e)}), 500


# ========== Hero Images API ==========

@app.get("/api/hero-images")
def get_hero_images():
    """Get all active hero images (public endpoint)."""
    with mysql_engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, alt_text, display_order, link_url
                FROM hero_carousel
                WHERE is_active = 1
                ORDER BY display_order ASC
                """
            )
        ).mappings().all()

    # Build URL pointing to the image data endpoint
    images = []
    for row in rows:
        images.append({
            "id": row["id"],
            "url": f"/api/hero-images/{row['id']}/data",
            "alt_text": row["alt_text"],
            "display_order": row["display_order"],
            "link_url": row["link_url"]
        })
    return jsonify({"success": True, "images": images})


@app.get("/api/hero-images/<int:image_id>/data")
def get_hero_image_data(image_id: int):
    """Serve hero image binary data."""
    from urllib.parse import quote

    with mysql_engine.begin() as conn:
        row = conn.execute(
            text("SELECT image_data, content_type, filename FROM hero_carousel WHERE id = :id AND is_active = 1"),
            {"id": image_id}
        ).mappings().first()

    if not row:
        abort(404)

    # URL encode filename to handle non-ASCII characters
    encoded_filename = quote(row["filename"])

    return Response(
        row["image_data"],
        mimetype=row["content_type"],
        headers={
            "Cache-Control": "public, max-age=86400",
            "Content-Disposition": f"inline; filename*=UTF-8''{encoded_filename}"
        }
    )


@app.get("/api/admin/hero-images")
@admin_required
def admin_get_hero_images():
    """Get all hero images (admin endpoint)."""
    with mysql_engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT id, filename, content_type, alt_text,
                       display_order, is_active, link_url,
                       created_at, updated_at
                FROM hero_carousel
                ORDER BY display_order ASC
                """
            )
        ).mappings().all()

    # Build URL pointing to the image data endpoint
    images = []
    for row in rows:
        images.append({
            "id": row["id"],
            "url": f"/api/hero-images/{row['id']}/data",
            "filename": row["filename"],
            "alt_text": row["alt_text"],
            "display_order": row["display_order"],
            "is_active": row["is_active"],
            "link_url": row["link_url"],
            "created_at": row["created_at"],
            "updated_at": row["updated_at"]
        })
    return jsonify({"success": True, "images": images})


@app.post("/api/admin/hero-images")
@admin_required
@csrf_protect
@limiter.limit("10 per hour")
def admin_upload_hero_image():
    """Upload a new hero image (stores in database)."""
    if "file" not in request.files:
        return jsonify({"success": False, "error": "未提供檔案"}), 400

    file = request.files["file"]

    # Use comprehensive file validator
    try:
        file_data = validate_image_upload(file)
    except FileValidationError as e:
        logger.warning(f"File validation failed: {e}")
        return jsonify({"success": False, "error": str(e)}), 400

    # Get alt_text from form
    alt_text = request.form.get("alt_text", "")

    # Get link_url from form and validate
    link_url = request.form.get("link_url", "")
    is_valid, error_msg = validate_url(link_url)
    if not is_valid:
        return jsonify({"success": False, "error": error_msg}), 400

    # Get next display order and insert into database
    with mysql_engine.begin() as conn:
        result = conn.execute(
            text("SELECT COALESCE(MAX(display_order), -1) + 1 as next_order FROM hero_carousel")
        ).mappings().first()
        next_order = result["next_order"] if result else 0

        # Check if we already have 8 images
        count_result = conn.execute(
            text("SELECT COUNT(*) as count FROM hero_carousel")
        ).mappings().first()
        if count_result and count_result["count"] >= 8:
            return jsonify({"success": False, "error": "最多只能上傳 8 張圖片"}), 400

        # Insert image data into database
        now = utcnow()
        conn.execute(
            text(
                """
                INSERT INTO hero_carousel (filename, content_type, image_data, alt_text, link_url, display_order, created_at, updated_at)
                VALUES (:filename, :content_type, :image_data, :alt_text, :link_url, :order, :now, :now)
                """
            ),
            {
                "filename": file.filename,
                "content_type": file.content_type,
                "image_data": file_data,
                "alt_text": alt_text,
                "link_url": link_url.strip() if link_url else None,
                "order": next_order,
                "now": now,
            },
        )

        # Get the inserted image
        inserted = conn.execute(
            text("SELECT id, alt_text, display_order, link_url FROM hero_carousel WHERE display_order = :order"),
            {"order": next_order}
        ).mappings().first()

    if inserted:
        log_admin_action('upload', 'hero_image', inserted['id'], {
            'filename': file.filename,
            'size': len(file_data)
        })

        return jsonify({
            "success": True,
            "image": {
                "id": inserted["id"],
                "url": f"/api/hero-images/{inserted['id']}/data",
                "alt_text": inserted["alt_text"],
                "display_order": inserted["display_order"],
                "link_url": inserted["link_url"]
            }
        })
    return jsonify({"success": False, "error": "上傳失敗"}), 500


@app.delete("/api/admin/hero-images/<int:image_id>")
@admin_required
@csrf_protect
def admin_delete_hero_image(image_id: int):
    """Delete a hero image from database."""
    with mysql_engine.begin() as conn:
        # Get image info for logging
        image = conn.execute(
            text("SELECT id, filename FROM hero_carousel WHERE id = :id"),
            {"id": image_id}
        ).mappings().first()

        if not image:
            return jsonify({"success": False, "error": "圖片不存在"}), 404

        # Delete from database
        conn.execute(
            text("DELETE FROM hero_carousel WHERE id = :id"),
            {"id": image_id}
        )

        log_admin_action('delete', 'hero_image', image_id, {'filename': image['filename']})

    return jsonify({"success": True, "message": "已刪除"})


@app.put("/api/admin/hero-images/reorder")
@admin_required
@csrf_protect
def admin_reorder_hero_images():
    """Reorder hero images."""
    data = request.get_json() or {}
    order = data.get("order", [])  # List of image IDs in new order

    if not isinstance(order, list):
        return jsonify({"success": False, "error": "order 必須是陣列"}), 400

    with mysql_engine.begin() as conn:
        for idx, image_id in enumerate(order):
            conn.execute(
                text(
                    """
                    UPDATE hero_carousel
                    SET display_order = :order, updated_at = :now
                    WHERE id = :id
                    """
                ),
                {"order": idx, "id": image_id, "now": utcnow()}
            )

    return jsonify({"success": True, "message": "排序已更新"})


@app.put("/api/admin/hero-images/<int:image_id>")
@admin_required
@csrf_protect
def admin_update_hero_image(image_id: int):
    """Update hero image metadata.

    Security: Uses field mapping dictionary to prevent SQL injection.
    All field names are predefined and never constructed from user input.
    """
    data = request.get_json() or {}

    # Field mapping dictionary: defines allowed fields and their SQL fragments
    # This eliminates the need for string concatenation in SQL construction
    FIELD_MAPPING = {
        'alt_text': 'alt_text = :alt_text',
        'is_active': 'is_active = :is_active',
        'link_url': 'link_url = :link_url',
    }

    updates = []
    params = {"id": image_id, "now": utcnow()}

    # Validate and build update statements
    for field, value in data.items():
        # Reject any fields not in the mapping
        if field not in FIELD_MAPPING:
            return jsonify({
                "success": False,
                "error": f"不允許更新的欄位: {field}"
            }), 400

        # Field-specific validation and processing
        if field == "is_active":
            params[field] = 1 if value else 0
        elif field == "link_url":
            # Validate URL
            is_valid, error_msg = validate_url(value)
            if not is_valid:
                return jsonify({"success": False, "error": error_msg}), 400
            params[field] = value.strip() if value else None
        else:
            params[field] = value

        # Use predefined SQL fragment from mapping
        updates.append(FIELD_MAPPING[field])

    if not updates:
        return jsonify({"success": False, "error": "沒有要更新的欄位"}), 400

    # Always update timestamp
    updates.append("updated_at = :now")

    # Build and execute SQL - safe because all field names come from FIELD_MAPPING
    sql = "UPDATE hero_carousel SET " + ", ".join(updates) + " WHERE id = :id"

    with mysql_engine.begin() as conn:
        result = conn.execute(text(sql), params)

        if result.rowcount == 0:
            return jsonify({"success": False, "error": "圖片不存在"}), 404

        log_admin_action('update', 'hero_image', image_id, {'updates': list(data.keys())})

    return jsonify({"success": True, "message": "已更新"})


def ensure_chat_session(session_id: Optional[str] = None, member_id: Optional[int] = None) -> str:
    """Return an existing chat session id or create a new one."""
    chat_session_id = session_id or uuid.uuid4().hex
    now = utcnow()
    with mysql_engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO chat_sessions (id, member_id, created_at, updated_at)
                VALUES (:id, :member_id, :now, :now)
                ON DUPLICATE KEY UPDATE updated_at = :now
                """
            ),
            {"id": chat_session_id, "member_id": member_id, "now": now},
        )
    return chat_session_id


def save_chat_message(session_id: str, role: str, content: str, template_id: Optional[str] = None) -> None:
    """Persist a chat message for a given session."""
    with mysql_engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO chat_messages (session_id, role, content, template_id, created_at)
                VALUES (:sid, :role, :content, :template_id, :created_at)
                """
            ),
            {
                "sid": session_id,
                "role": role,
                "content": content,
                "template_id": template_id,
                "created_at": utcnow(),
            },
        )
        conn.execute(
            text(
                """
                UPDATE chat_sessions
                SET updated_at = :updated_at
                WHERE id = :sid
                """
            ),
            {"sid": session_id, "updated_at": utcnow()},
        )


def fetch_chat_history(session_id: str, limit: int = 12) -> List[Dict[str, Any]]:
    """Fetch the most recent chat history for the session in chronological order."""
    # Strict validation to prevent SQL injection
    if not isinstance(limit, int):
        raise ValueError("limit must be an integer")
    if limit <= 0:
        limit = 1
    if limit > 100:  # Maximum safety limit
        limit = 100

    # Safe to use in query after validation
    query = text(
        f"""
        SELECT role, content
        FROM chat_messages
        WHERE session_id = :sid
        ORDER BY created_at DESC
        LIMIT {int(limit)}
        """
    )

    with mysql_engine.begin() as conn:
        rows = conn.execute(query, {"sid": session_id}).mappings().all()

    # Reverse to chronological order
    return [dict(row) for row in reversed(rows)]


def build_chat_history(history: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    """Prepare chat history for OpenAI API (without system/user prompts - handled by RAG)."""
    messages: List[Dict[str, str]] = []
    for item in history:
        role = item.get("role")
        content = (item.get("content") or "").strip()
        if role in {"user", "assistant"} and content:
            messages.append({"role": role, "content": content})
    return messages


def format_sse(payload: Dict[str, Any]) -> str:
    """Serialize a Python dictionary into a Server-Sent Events data frame."""
    return f"data: {json.dumps(payload, ensure_ascii=False)}\n\n"


# Regex pattern to remove OpenAI file search citation markers (e.g., fileciteturn0file5turn0file12)
_CITATION_PATTERN = re.compile(r"fileciteturn\d+file\d+(?:turn\d+file\d+)*")

# Pre-compiled emoji pattern: matches emoji ranges while preserving CJK characters (U+2E80-U+9FFF)
_EMOJI_PATTERN = re.compile(
    "["
    "\U0001F600-\U0001F64F"  # emoticons
    "\U0001F300-\U0001F5FF"  # symbols & pictographs
    "\U0001F680-\U0001F6FF"  # transport & map
    "\U0001F1E0-\U0001F1FF"  # flags
    "\U0001F900-\U0001F9FF"  # supplemental symbols
    "\U0001FA00-\U0001FA6F"  # chess symbols
    "\U0001FA70-\U0001FAFF"  # symbols extended-A
    "\U0000FE00-\U0000FE0F"  # variation selectors
    "\U0000200D"             # zero width joiner
    "\U00002B50"             # star
    "\U000024C2-\U000024FF"  # enclosed alphanumerics (narrow)
    "\U00002702-\U000027B0"  # dingbats
    "\U00002600-\U000026FF"  # misc symbols
    "]+",
    flags=re.UNICODE,
)

_MULTI_SPACE_PATTERN = re.compile(r'  +')


def strip_citations(text: str) -> str:
    """Remove OpenAI file search citation markers from text."""
    return _CITATION_PATTERN.sub("", text)


def _strip_emojis(text: str) -> str:
    """移除文字中的 emoji 和表情符號，保留中文字元"""
    result = _EMOJI_PATTERN.sub("", text)
    return _MULTI_SPACE_PATTERN.sub(" ", result).strip()


def fix_activity_list_format(text: str) -> str:
    """
    修正活動列表格式，確保編號和活動名稱在同一行。

    處理以下情況：
    1. 編號和活動名稱分行 → 合併到同一行
    2. 移除不必要的粗體標記（新格式不使用粗體）
    3. 移除活動名稱中的 emoji
    """
    lines = text.split('\n')
    result_lines = []
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # 模式 1：編號獨佔一行（可能帶粗體或「活動名稱：」）
        # 匹配：1. 或 **1.** 或 1. 活動名稱：（後面沒有實際標題）
        number_only = re.match(r'^(?:\*\*)?(\d+)\.(?:\*\*)?\s*(?:活動名稱：)?\s*$', line)
        if number_only and i + 1 < len(lines):
            number = number_only.group(1)
            next_line = lines[i + 1].strip()
            # 移除下一行的粗體標記和「活動名稱：」前綴
            next_line = re.sub(r'^\*\*|\*\*$', '', next_line).strip()
            next_line = re.sub(r'^活動名稱[：:]', '', next_line).strip()
            next_line = _strip_emojis(next_line)
            if next_line:
                result_lines.append(f'{number}. 活動名稱：{next_line}')
                i += 2
                continue

        # 模式 2：有編號+標題但使用粗體格式 → 移除粗體，改為新格式
        # **1. 標題** 或 **1. 活動名稱：標題**
        bold_pattern = re.match(r'^\*\*(\d+)\.\s*(?:活動名稱[：:])?\s*(.+?)\*\*$', line)
        if bold_pattern:
            number = bold_pattern.group(1)
            title = _strip_emojis(bold_pattern.group(2).strip())
            result_lines.append(f'{number}. 活動名稱：{title}')
            i += 1
            continue

        # 模式 3：有編號+標題但沒有「活動名稱：」前綴
        # 1. 青年創業講座（不帶粗體）
        plain_pattern = re.match(r'^(\d+)\.\s+(?!活動[名日]|報名|貼文)(.+)$', line)
        if plain_pattern:
            number = plain_pattern.group(1)
            title = plain_pattern.group(2).strip()
            title = re.sub(r'^\*\*|\*\*$', '', title).strip()
            title = _strip_emojis(title)
            result_lines.append(f'{number}. 活動名稱：{title}')
            i += 1
            continue

        # 模式 4：已有「活動名稱：」格式但含有 emoji
        name_pattern = re.match(r'^(\d+)\.\s*活動名稱[：:]\s*(.+)$', line)
        if name_pattern:
            number = name_pattern.group(1)
            title = _strip_emojis(name_pattern.group(2).strip())
            result_lines.append(f'{number}. 活動名稱：{title}')
            i += 1
            continue

        # 其他行保持不變
        result_lines.append(lines[i])
        i += 1

    return '\n'.join(result_lines)


@app.post("/api/chat")
@app.post("/chat")
@csrf_protect
@validate_message_input
@limiter.limit("30 per minute")
def api_chat():
    if not request.is_json:
        return jsonify({"error": "Payload must be JSON."}), 400

    payload = request.get_json(force=True) or {}
    message = (payload.get("message") or "").strip()
    requested_session = payload.get("session_id")
    template_id = payload.get("template_id")  # Extract template_id from payload

    if not message:
        return jsonify({"error": "message is required"}), 400

    # Get logged-in user's member_id from session
    member_id = session.get("user", {}).get("member_id")

    session_id = ensure_chat_session(
        requested_session if isinstance(requested_session, str) else None,
        member_id=member_id
    )
    history = fetch_chat_history(session_id)

    # Persist the user's message before streaming.
    save_chat_message(session_id, "user", message, template_id)

    client = OPENAI_CLIENT
    rag_store = get_rag_store_name()

    def generate():
        logger.info("Streaming response for session %s", session_id)
        yield format_sse({"type": "session", "content": "", "session_id": session_id})

        if client is None or rag_store is None:
            assistant_text = (
                "無法連接 OpenAI 服務。請檢查伺服器設定。\n\n"
                f"待發送訊息：{message}\n"
                "請確認 OPENAI_API_KEY 已設定後再試。"
            )
            save_chat_message(session_id, "assistant", assistant_text)
            yield format_sse(
                {"type": "text", "content": assistant_text, "session_id": session_id}
            )
            yield format_sse({"type": "end", "content": "", "session_id": session_id})
            return

        accumulated: List[str] = []
        sources: List[Dict[str, Any]] = []

        try:
            chat_history = build_chat_history(history)

            for chunk in generate_with_rag_stream(
                query=message,
                system_prompt=SYSTEM_PROMPT,
                chat_history=chat_history,
                model=OPENAI_MODEL
            ):
                if chunk["type"] == "text":
                    delta = chunk["content"]
                    if delta:
                        # Remove citation markers before sending to client
                        clean_delta = strip_citations(delta)
                        accumulated.append(clean_delta)
                        if clean_delta:
                            yield format_sse(
                                {
                                    "type": "text",
                                    "content": clean_delta,
                                    "session_id": session_id,
                                }
                            )
                elif chunk["type"] == "sources":
                    sources = chunk["content"]
                elif chunk["type"] == "end":
                    pass

        except Exception as e:
            from openai import RateLimitError, APITimeoutError, OpenAIError
            from sqlalchemy.exc import SQLAlchemyError

            error_message = "產生回覆時發生問題，請稍後再試或聯繫我們的服務人員。"

            if isinstance(e, RateLimitError):
                logger.warning(f"OpenAI rate limit hit for session {session_id}")
                error_message = "服務暫時過載，請稍後重試"
            elif isinstance(e, APITimeoutError):
                logger.warning(f"OpenAI timeout for session {session_id}")
                error_message = "AI 服務響應超時，請重試"
            elif isinstance(e, OpenAIError):
                logger.error(f"OpenAI API error: {e}", exc_info=True)
                error_message = "AI 服務暫時不可用"
            elif isinstance(e, SQLAlchemyError):
                logger.error(f"Database error in chat: {e}", exc_info=True)
                error_message = "數據庫錯誤，請重試"
            elif isinstance(e, ValueError):
                logger.warning(f"Invalid input: {e}")
                error_message = str(e)
            else:
                logger.critical(f"Unexpected error in chat: {e}", exc_info=True)

            save_chat_message(session_id, "assistant", error_message)
            yield format_sse(
                {"type": "error", "content": error_message, "session_id": session_id}
            )
            yield format_sse({"type": "end", "content": "", "session_id": session_id})
            return

        full_text = "".join(accumulated).strip()

        # 修正活動列表格式（確保編號和標題在同一行且使用粗體）
        if full_text:
            full_text = fix_activity_list_format(full_text)
            save_chat_message(session_id, "assistant", full_text)
        else:
            fallback_text = "抱歉，我目前無法回覆。請重新描述您的問題或聯繫我們的服務人員。"
            full_text = fallback_text
            save_chat_message(session_id, "assistant", fallback_text)
            yield format_sse(
                {"type": "text", "content": fallback_text, "session_id": session_id}
            )

        # Yield sources if available
        if sources:
            yield format_sse(
                {"type": "sources", "content": sources, "session_id": session_id}
            )

        # 送出格式化後的完整文字，讓前端替換串流累積的未格式化版本
        yield format_sse({"type": "end", "content": full_text, "session_id": session_id})

    response = Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
    )
    response.headers["Cache-Control"] = "no-cache"
    response.headers["X-Accel-Buffering"] = "no"
    return response


def fetchall(sql: str, params: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    with mysql_engine.begin() as conn:
        return [
            dict(row)
            for row in conn.execute(text(sql), params or {}).mappings().all()
        ]


def fetchone(sql: str, params: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
    with mysql_engine.begin() as conn:
        result = conn.execute(text(sql), params or {}).mappings().first()
        return dict(result) if result else None


def execute(sql: str, params: Optional[Dict[str, Any]] = None) -> None:
    with mysql_engine.begin() as conn:
        conn.execute(text(sql), params or {})


def _clean(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    trimmed = str(value).strip()
    return trimmed or None


def upsert_member(
    external_id: Optional[str],
    display_name: Optional[str] = None,
    avatar_url: Optional[str] = None,
    gender: Optional[str] = None,
    birthday: Optional[str] = None,
    email: Optional[str] = None,
    phone: Optional[str] = None,
    source: Optional[str] = "form",
) -> Optional[int]:
    """Create or update a member record identified by an external id."""
    external_id = _clean(external_id)
    if not external_id:
        return None

    now = utcnow()
    data = {
        "display_name": _clean(display_name),
        "avatar_url": _clean(avatar_url),
        "gender": _clean(gender),
        "birthday": _clean(birthday),
        "email": _clean(email),
        "phone": _clean(phone),
        "source": source or "form",
        "updated_at": now,
        "last_interaction_at": now,
    }

    with mysql_engine.begin() as conn:
        existing = conn.execute(
            text("SELECT id FROM members WHERE external_id = :ext"),
            {"ext": external_id},
        ).scalar()
        if existing:
            conn.execute(
                text(
                    """
                    UPDATE members
                       SET display_name=:display_name,
                           avatar_url=:avatar_url,
                           gender=:gender,
                           birthday=:birthday,
                           email=:email,
                           phone=:phone,
                           source=:source,
                           updated_at=:updated_at,
                           last_interaction_at=:last_interaction_at
                     WHERE id=:member_id
                    """
                ),
                {**data, "member_id": existing},
            )
            return int(existing)

        insert_params = {
            "external_id": external_id,
            **data,
            "created_at": now,
        }
        result = conn.execute(
            text(
                """
                INSERT INTO members (
                    external_id,
                    display_name,
                    avatar_url,
                    gender,
                    birthday,
                    email,
                    phone,
                    source,
                    created_at,
                    updated_at,
                    last_interaction_at
                ) VALUES (
                    :external_id,
                    :display_name,
                    :avatar_url,
                    :gender,
                    :birthday,
                    :email,
                    :phone,
                    :source,
                    :created_at,
                    :updated_at,
                    :last_interaction_at
                )
                """
            ),
            insert_params,
        )
        member_id = result.lastrowid
    return int(member_id) if member_id is not None else None



# Survey-related functions and templates removed



@app.get("/")
def index():
    """Serve the index.html file from dist directory in production, or BASE_DIR in development"""
    # 优先从 dist 目录提供（生产环境），否则从 BASE_DIR（开发环境）
    dist_index = os.path.join(DIST_DIR, "index.html")
    if os.path.exists(dist_index):
        return send_from_directory(DIST_DIR, "index.html")
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/<path:path>")
def serve_static(path: str):
    """Serve static files from dist directory"""
    # 检查是否是 API 路由，如果是则跳过
    if path.startswith("api/") or path.startswith("__"):
        abort(404)
    
    # 优先从 dist 目录提供静态文件
    dist_path = os.path.join(DIST_DIR, path)
    if os.path.exists(dist_path) and os.path.isfile(dist_path):
        return send_from_directory(DIST_DIR, path)
    
    # 如果文件不存在，返回 index.html（用于 SPA 路由）
    dist_index = os.path.join(DIST_DIR, "index.html")
    if os.path.exists(dist_index):
        return send_from_directory(DIST_DIR, "index.html")
    
    abort(404)


@app.get("/health")
def health() -> tuple[Dict[str, Any], int]:
    """Health check endpoint with database connection test."""
    health_status: Dict[str, Any] = {
        "status": "healthy",
        "database": "unknown",
        "timestamp": datetime.datetime.now(timezone.utc).isoformat(),
    }

    try:
        with mysql_engine.connect() as conn:
            conn.execute(text("SELECT 1"))
            conn.commit()
        health_status["database"] = "connected"
    except Exception as e:
        logger.error("Health check failed: MySQL connection error: %s", e)
        health_status["status"] = "unhealthy"
        health_status["database"] = "disconnected"
        health_status["error"] = str(e)
        return health_status, 503

    return health_status, 200


# ==================== OAuth Routes ====================

@app.get("/auth/config")
def api_auth_config():
    """Return OAuth configuration for frontend (without secrets)."""
    return jsonify({
        "google": {
            "enabled": bool(GOOGLE_CLIENT_ID),
            "client_id": GOOGLE_CLIENT_ID,
            "redirect_uri": GOOGLE_REDIRECT_URI
        },
        "line": {
            "enabled": bool(LINE_CHANNEL_ID),
            "channel_id": LINE_CHANNEL_ID,
            "redirect_uri": LINE_REDIRECT_URI
        },
        "facebook": {
            "enabled": bool(FACEBOOK_APP_ID),
            "app_id": FACEBOOK_APP_ID,
            "redirect_uri": FACEBOOK_REDIRECT_URI
        }
    })


# OAuth state configuration
OAUTH_VALID_PROVIDERS = {"google", "line", "facebook"}
OAUTH_STATE_EXPIRY_SECONDS = 900  # 15 minutes


@app.post("/api/auth/state/<provider>")
def generate_oauth_state(provider: str):
    """Generate and store a cryptographically secure OAuth state parameter.

    The state parameter prevents CSRF attacks during OAuth flows by ensuring
    the callback request originated from this application.

    Args:
        provider: OAuth provider name (google, line, or facebook)

    Returns:
        JSON response containing the generated state token
    """
    if provider not in OAUTH_VALID_PROVIDERS:
        return jsonify({"error": "Invalid provider"}), 400

    state = secrets.token_urlsafe(32)
    session_key = f"oauth_state_{provider}"

    session[session_key] = {
        "state": state,
        "created_at": utcnow().isoformat()
    }
    session.permanent = False  # Temporary session for OAuth flow

    logger.info(f"Generated OAuth state for provider: {provider}")
    return jsonify({"state": state})


def validate_oauth_state(provider: str, received_state: Optional[str]) -> bool:
    """Validate OAuth state parameter to prevent CSRF attacks.

    Performs three security checks:
    1. State parameter presence and match (constant-time comparison)
    2. Expiration check (must be within OAUTH_STATE_EXPIRY_SECONDS)
    3. One-time use (state is cleared after successful validation)

    Args:
        provider: OAuth provider name (google, line, or facebook)
        received_state: State parameter from the OAuth callback

    Returns:
        True if state is valid and not expired, False otherwise
    """
    if not received_state:
        logger.warning(f"OAuth callback missing state parameter: {provider}")
        return False

    session_key = f"oauth_state_{provider}"
    stored_data = session.get(session_key)

    if not stored_data:
        logger.warning(f"No stored state found for provider: {provider}")
        return False

    stored_state = stored_data.get("state")
    created_at_str = stored_data.get("created_at")

    # Security: Use constant-time comparison to prevent timing attacks
    if not secrets.compare_digest(received_state, stored_state):
        logger.warning(f"OAuth state mismatch for provider: {provider}")
        return False

    # Validate expiration timestamp
    if not created_at_str:
        logger.error(f"Missing created_at timestamp for provider: {provider}")
        return False

    try:
        created_at = datetime.datetime.fromisoformat(created_at_str)
        # Make timezone-aware for comparison (utcnow returns naive UTC datetime)
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=datetime.timezone.utc)
        age_seconds = (datetime.datetime.now(datetime.timezone.utc) - created_at).total_seconds()

        if age_seconds > OAUTH_STATE_EXPIRY_SECONDS:
            logger.warning(f"OAuth state expired for provider: {provider} (age: {age_seconds:.0f}s)")
            return False
    except (ValueError, TypeError) as e:
        logger.error(f"Invalid created_at timestamp for provider: {provider}, error: {e}")
        return False

    # Security: Clear state after validation (one-time use prevents replay attacks)
    session.pop(session_key, None)

    logger.info(f"OAuth state validated successfully for provider: {provider}")
    return True


@app.get("/auth/google/callback")
def auth_google_callback():
    """Handle Google OAuth callback."""
    code = request.args.get("code")
    error = request.args.get("error")
    state = request.args.get("state")

    # Validate state parameter to prevent CSRF attacks
    if not validate_oauth_state("google", state):
        logger.error("Google OAuth: Invalid or missing state parameter")
        return redirect("/?error=oauth_csrf_validation_failed")

    if error:
        logger.error(f"Google OAuth error: {error}")
        return redirect("/?error=google_auth_failed")

    if not code:
        return redirect("/?error=no_code")

    try:
        # Exchange code for token
        token_response = http_requests.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": code,
                "client_id": GOOGLE_CLIENT_ID,
                "client_secret": GOOGLE_CLIENT_SECRET,
                "redirect_uri": GOOGLE_REDIRECT_URI,
                "grant_type": "authorization_code"
            }
        )
        token_data = token_response.json()
        access_token = token_data.get("access_token")

        if not access_token:
            logger.error(f"Google token exchange failed: {token_data}")
            return redirect("/?error=google_token_exchange_failed")

        # Get user info
        user_response = http_requests.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {access_token}"}
        )
        user_info = user_response.json()

        # Upsert member and store in session
        member_id = upsert_member(
            external_id=f"google_{user_info['id']}",
            display_name=user_info.get("name"),
            avatar_url=user_info.get("picture"),
            email=user_info.get("email"),
            source="google"
        )

        # Regenerate session ID to prevent session fixation attack
        session.clear()
        session.modified = True

        session["user"] = {
            "member_id": member_id,
            "provider": "google",
            "external_id": f"google_{user_info['id']}",
            "email": user_info.get("email"),
            "name": user_info.get("name")
        }
        session.permanent = True

        return redirect("/?login=success")

    except Exception as e:
        logger.exception("Google OAuth token exchange failed")
        return redirect("/?error=google_token_exchange_failed")


@app.get("/auth/line/callback")
def auth_line_callback():
    """Handle LINE OAuth callback."""
    code = request.args.get("code")
    error = request.args.get("error")
    state = request.args.get("state")

    # Validate state parameter to prevent CSRF attacks
    if not validate_oauth_state("line", state):
        logger.error("LINE OAuth: Invalid or missing state parameter")
        return redirect("/?error=oauth_csrf_validation_failed")

    if error:
        logger.error(f"LINE OAuth error: {error}")
        return redirect("/?error=line_auth_failed")

    if not code:
        return redirect("/?error=no_code")

    try:
        # Exchange code for token (LINE requires form-urlencoded)
        token_response = http_requests.post(
            "https://api.line.me/oauth2/v2.1/token",
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": LINE_REDIRECT_URI,
                "client_id": LINE_CHANNEL_ID,
                "client_secret": LINE_CHANNEL_SECRET
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"}
        )
        token_data = token_response.json()
        access_token = token_data.get("access_token")

        if not access_token:
            logger.error(f"LINE token exchange failed: {token_data}")
            return redirect("/?error=line_token_exchange_failed")

        # Get user profile
        profile_response = http_requests.get(
            "https://api.line.me/v2/profile",
            headers={"Authorization": f"Bearer {access_token}"}
        )
        profile = profile_response.json()

        # Upsert member and store in session
        member_id = upsert_member(
            external_id=f"line_{profile['userId']}",
            display_name=profile.get("displayName"),
            avatar_url=profile.get("pictureUrl"),
            source="line"
        )

        # Regenerate session ID to prevent session fixation attack
        session.clear()
        session.modified = True

        session["user"] = {
            "member_id": member_id,
            "provider": "line",
            "external_id": f"line_{profile['userId']}",
            "name": profile.get("displayName")
        }
        session.permanent = True

        logger.info(f"[LINE Login] User logged in: {profile.get('displayName')} (member_id: {member_id}, external_id: line_{profile['userId']})")

        return redirect("/?login=success")

    except Exception as e:
        logger.exception("LINE OAuth token exchange failed")
        return redirect("/?error=line_token_exchange_failed")


@app.get("/auth/facebook/callback")
def auth_facebook_callback():
    """Handle Facebook OAuth callback."""
    code = request.args.get("code")
    error = request.args.get("error")
    state = request.args.get("state")

    # Validate state parameter to prevent CSRF attacks
    if not validate_oauth_state("facebook", state):
        logger.error("Facebook OAuth: Invalid or missing state parameter")
        return redirect("/?error=oauth_csrf_validation_failed")

    if error:
        logger.error(f"Facebook OAuth error: {error}")
        return redirect("/?error=facebook_auth_failed")

    if not code:
        return redirect("/?error=no_code")

    try:
        # Exchange code for token
        token_response = http_requests.get(
            "https://graph.facebook.com/v18.0/oauth/access_token",
            params={
                "client_id": FACEBOOK_APP_ID,
                "client_secret": FACEBOOK_APP_SECRET,
                "redirect_uri": FACEBOOK_REDIRECT_URI,
                "code": code
            }
        )
        token_data = token_response.json()
        access_token = token_data.get("access_token")

        if not access_token:
            logger.error(f"Facebook token exchange failed: {token_data}")
            return redirect("/?error=facebook_token_exchange_failed")

        # Get user profile
        profile_response = http_requests.get(
            "https://graph.facebook.com/me",
            params={
                "fields": "id,name,email,picture.type(large)",
                "access_token": access_token
            }
        )
        profile = profile_response.json()

        picture_url = None
        if profile.get("picture") and profile["picture"].get("data"):
            picture_url = profile["picture"]["data"].get("url")

        # Upsert member and store in session
        member_id = upsert_member(
            external_id=f"facebook_{profile['id']}",
            display_name=profile.get("name"),
            avatar_url=picture_url,
            email=profile.get("email"),
            source="facebook"
        )

        # Regenerate session ID to prevent session fixation attack
        session.clear()
        session.modified = True

        session["user"] = {
            "member_id": member_id,
            "provider": "facebook",
            "external_id": f"facebook_{profile['id']}",
            "email": profile.get("email"),
            "name": profile.get("name")
        }
        session.permanent = True

        return redirect("/?login=success")

    except Exception as e:
        logger.exception("Facebook OAuth token exchange failed")
        return redirect("/?error=facebook_token_exchange_failed")


@app.get("/api/user")
def api_get_user():
    """Return current authenticated user or guest status.

    Returns 200 with success=false for unauthenticated users to avoid
    console warnings while maintaining clear authentication state.
    """
    # Debug: 記錄 session 資訊
    from flask import request as flask_request
    logger.info(f"[/api/user] Session ID cookie: {flask_request.cookies.get('session', 'N/A')[:20] if flask_request.cookies.get('session') else 'None'}...")
    logger.info(f"[/api/user] Session data: {dict(session) if session else 'Empty'}")
    if "user" in session:
        user_data = session["user"].copy()
        # 從資料庫讀取頭像 (避免 session cookie 過大導致 431 錯誤)
        if user_data.get("member_id"):
            member = fetchone(
                "SELECT avatar_url FROM members WHERE id = :id",
                {"id": user_data["member_id"]}
            )
            if member:
                user_data["picture"] = member.get("avatar_url")
        response = jsonify({
            "success": True,
            "user": user_data
        })
        return _no_store(response)
    # Return 200 with success=false for unauthenticated users
    # This avoids console warnings while clearly indicating no user is logged in
    response = jsonify({
        "success": False,
        "message": "Not authenticated"
    })
    return _no_store(response)


@app.post("/api/logout")
def api_logout():
    """Destroy session and logout user."""
    session.clear()
    return jsonify({
        "success": True,
        "message": "Logged out successfully"
    })


@app.get(f"{ASSET_ROUTE_PREFIX}/<path:filename>")
def serve_uploads(filename: str):
    return send_from_directory(ASSET_LOCAL_DIR, filename, conditional=True)



if __name__ == "__main__":
    port = int(os.getenv("PORT", "8300"))
    debug_mode = os.getenv("FLASK_DEBUG", "0") in {"1", "true", "True"}
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
